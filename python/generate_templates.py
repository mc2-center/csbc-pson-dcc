"""Generate manifest templates."""

import os
import argparse
import getpass

import pandas as pd
import synapseclient
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font


def login():
    """Log into Synapse. If env variables not found, prompt user.

    Returns:
        syn: Synapse object
    """
    try:
        syn = synapseclient.login(
            os.getenv('SYN_USERNAME'),
            apiKey=os.getenv('SYN_APIKEY'),
            silent=True)
    except synapseclient.core.exceptions.SynapseNoCredentialsError:
        print("Credentials not found; please manually provide your",
              "Synapse username and password.")
        username = input("Synapse username: ")
        password = getpass.getpass("Synapse password: ")
        syn = synapseclient.login(username, password, silent=True)
    return syn


def download_table(syn, table_id, col_names="*", params=None):
    """Download and return Synapse table as a dataframe."""
    query = f"SELECT {col_names} FROM {table_id}"

    if params:
        query += f" WHERE {params}"
    return syn.tableQuery(query).asDataFrame().fillna("").drop_duplicates()


def generate_latest_template(old_wb, new_wb, cv_terms):
    """Update the manifest template with new CV terms."""
    wb = load_workbook(old_wb)

    # Replace the worksheet with older CV with the latest terms.
    old_terms = wb['standard_terms']
    wb.remove(old_terms)
    new_terms = wb.create_sheet("standard_terms")
    for row in dataframe_to_rows(cv_terms, index=False, header=True):
        new_terms.append(row)

    # Style the worksheet.
    ft = Font(bold=True)
    new_terms["A1"].font = ft
    new_terms["B1"].font = ft
    new_terms["C1"].font = ft
    new_terms.column_dimensions['A'].width = 18
    new_terms.column_dimensions['B'].width = 60
    new_terms.column_dimensions['C'].width = 12
    new_terms.protection.sheet = True

    # Save the updated workbook.
    wb.save(new_wb)


def main():
    syn = login()

    dataset = os.path.join("manifest_templates", "datasets_manifest.xlsx")
    file = os.path.join("manifest_templates", "files_manifest.xlsx")
    tool = os.path.join("manifest_templates", "tools_manifest.xlsx")

    cv_terms = download_table(
        syn, "syn26433610",
        "key, value, columnType",
        "key <> '' AND columnType <> '' ORDER BY key, value")

    generate_latest_template(
        dataset,
        os.path.join("output", "datasets_manifest.xlsx"),
        cv_terms
    )
    generate_latest_template(
        file,
        os.path.join("output", "files_manifest.xlsx"),
        cv_terms
    )
    generate_latest_template(
        tool,
        os.path.join("output", "tools_manifest.xlsx"),
        cv_terms
    )


if __name__ == "__main__":
    main()
